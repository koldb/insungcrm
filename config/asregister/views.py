from django.shortcuts import render, redirect, get_object_or_404, reverse
import sys

sys.path.append('..')
from isscm.decorators import login_required
from isscm.models import Product_Management
from . import models
from .models import ASsheet, ASUploadFile
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum
import datetime
import xlwt
import openpyxl
from django.http import HttpResponse, HttpResponseRedirect
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta
import mimetypes
import shutil



# Create your views here.

# 임시 메인페이지
def index(request):
    login_session = request.session.get('login_session')
    return render(request, 'isscm/index.html', {'login_session': login_session})


# AS 입력
@login_required
def as_insert(request):
    print('as 입력 도달')
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')
    user_phone = request.session.get('user_phone')
    print(user_name)
    print(user_phone)
    # render context로 넘길때 key:value 로 넘겨야 넘어가고 받아진다

    if request.method == 'GET':
        print('겟 도달')
        # render context로 넘길때 key:value 로 넘겨야 넘어가고 받아진다
        context = {'login_session': login_session, 'user_name': user_name, 'user_phone': user_phone}
        print('겟 끝나 나감')
        return render(request, 'assheet/as_insert.html', context)
    elif request.method == 'POST':
        print("입력 시작")
        insert = ASsheet()
        insert.rp_date = request.POST['rp_date']
        insert.cname = request.POST['cname']
        insert.cuser = request.POST['cuser']
        insert.cphone = request.POST['cphone']
        insert.product_name = request.POST['product_name']
        insert.memo = request.POST['memo']
        insert.serial = request.POST['serial']
        insert.site = request.POST['site']
        insert.symptom = request.POST['symptom']

        try:
            ex_pm = Product_Management.objects.exclude(status='폐기')
            pm_modify = get_object_or_404(ex_pm, serial=request.POST['serial'])
            pm_modify.product_name = request.POST['product_name']
            pm_modify.current_location = request.POST['cname']
            pm_modify.status = "AS"
            pm_modify.serial = request.POST['serial']

            pm_modify.save()
            print("pm 까지 수정 저장완료")
        except:
            print("수정 저장완료")
            insert.save()

        insert.save()

        print('입력 끝나 나감')
        # return render(request, 'assheet/as_insert.html', context)
        return redirect('asregister:as_list')


# AS 접수건 리스트
@login_required
def as_list(request):
    print("리스트 시작")
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')
    user_phone = request.session.get('user_phone')
    global search_sort
    global startdate
    global enddate
    if request.method == 'GET':
        if login_session == 'insung':
            sort = request.GET.get('sort', '')
            query = request.GET.get('q', '')
            if sort == 'rg_date':
                company_sheet = ASsheet.objects.all().order_by('-rg_date')
            elif sort == 'rp_date':
                company_sheet = ASsheet.objects.all().order_by('-rp_date')
            elif sort == 'end_date':
                company_sheet = ASsheet.objects.all().order_by('-end_date')
            elif sort == 'product_name':
                company_sheet = ASsheet.objects.all().order_by('-product_name', '-rg_date')
            elif sort == 'finish':
                company_sheet = ASsheet.objects.all().order_by('-finish', '-rg_date')
            elif sort == 'asaction':
                company_sheet = ASsheet.objects.all().order_by('-asaction', '-rg_date')
            elif sort == 'la_category':
                company_sheet = ASsheet.objects.all().order_by('-la_category', '-rg_date')
            elif sort == 'me_category':
                company_sheet = ASsheet.objects.all().order_by('-me_category', '-rg_date')
            elif sort == 'sm_category':
                company_sheet = ASsheet.objects.all().order_by('-sm_category', '-rg_date')
            elif sort == 'user_name':
                company_sheet = ASsheet.objects.all().order_by('-user_name', '-rg_date')
            elif sort == 'cname':
                company_sheet = ASsheet.objects.all().order_by('-cname', '-rg_date')
            elif sort == 'all':
                company_sheet = ASsheet.objects.all().order_by('-rg_date', 'finish')
            else:
                print("리스트 조회 겸 목록 조회")
                search_sort = request.GET.get('search_sort', '')
                startdate = request.GET.get('sdate', '')
                enddate = request.GET.get('edate', '')
                if search_sort == 'cname':
                    company_sheet = ASsheet.objects.all().filter(cname__icontains=query).order_by('-rg_date',
                                                                                                  'finish')
                elif search_sort == 'cuser':
                    company_sheet = ASsheet.objects.all().filter(cuser__icontains=query).order_by('-rg_date',
                                                                                                  'finish')
                elif search_sort == 'product_name':
                    company_sheet = ASsheet.objects.all().filter(product_name__icontains=query).order_by('-rg_date',
                                                                                                         'finish')
                elif search_sort == 'serial':
                    company_sheet = ASsheet.objects.all().filter(serial__icontains=query).order_by('-rg_date',
                                                                                                   'finish')
                elif search_sort == 'after_serial':
                    company_sheet = ASsheet.objects.all().filter(after_serial__icontains=query).order_by('-rg_date',
                                                                                                         'finish')
                elif search_sort == 'asaction':
                    company_sheet = ASsheet.objects.all().filter(asaction__icontains=query).order_by('-rg_date',
                                                                                                     'finish')
                elif search_sort == 'la_category':
                    company_sheet = ASsheet.objects.all().filter(la_category__icontains=query).order_by('-rg_date',
                                                                                                        'finish')
                elif search_sort == 'me_category':
                    company_sheet = ASsheet.objects.all().filter(me_category__icontains=query).order_by('-rg_date',
                                                                                                        'finish')
                elif search_sort == 'sm_category':
                    company_sheet = ASsheet.objects.all().filter(sm_category__icontains=query).order_by('-rg_date',
                                                                                                        'finish')
                elif search_sort == 'all':
                    company_sheet = ASsheet.objects.all().filter(
                        Q(product_name__icontains=query) | Q(memo__icontains=query) | Q(cname__icontains=query)
                        | Q(finish__icontains=query) | Q(cuser__icontains=query) | Q(cphone__icontains=query)
                        | Q(option__icontains=query) | Q(serial__icontains=query) | Q(
                            symptom__icontains=query)).order_by('-rg_date', 'finish')
                elif search_sort == 'rg_date':
                    e_date = datetime.datetime.strptime(enddate, '%Y-%m-%d') + datetime.timedelta(hours=23, minutes=59,
                                                                                                  seconds=59)
                    company_sheet = ASsheet.objects.all().filter(rg_date__gte=startdate, rg_date__lte=e_date).order_by(
                        '-rg_date', 'finish')
                elif search_sort == 'rp_date':
                    company_sheet = ASsheet.objects.all().filter(rp_date__range=[startdate, enddate]).order_by(
                        '-rg_date', 'finish')
                elif search_sort == 'end_date':
                    company_sheet = ASsheet.objects.all().filter(end_date__range=[startdate, enddate]).order_by(
                        '-rg_date', 'finish')
                else:
                    company_sheet = ASsheet.objects.all().order_by('-rg_date', 'finish')

            # 한달이상 미처리건 조회
            over_as = ASsheet.objects.filter(rg_date__lte=date.today() - relativedelta(months=1)).exclude(
                finish='종료').order_by('-rg_date')

            page = request.GET.get('page', '1')
            paginator = Paginator(company_sheet, 10)
            page_obj = paginator.get_page(page)
            upfile = ASUploadFile.objects.all()

            context = {'login_session': login_session, 'company_sheet': company_sheet, 'page_obj': page_obj,
                       'sort': sort, 'user_name': user_name, 'search_sort': search_sort,
                       'over_as': over_as, 'query': query, 'sdate': startdate, 'edate': enddate,
                       'user_phone': user_phone}

        else:
            sort = request.GET.get('sort', '')
            query = request.GET.get('q', '')
            if sort == 'rg_date':
                company_sheet = ASsheet.objects.filter(cname=login_session).order_by('-rg_date')
            elif sort == 'rp_date':
                company_sheet = ASsheet.objects.filter(cname=login_session).order_by('-rp_date')
            elif sort == 'product_name':
                company_sheet = ASsheet.objects.filter(cname=login_session).order_by('-product_name', '-rg_date')
            elif sort == 'finish':
                company_sheet = ASsheet.objects.filter(cname=login_session).order_by('-finish', '-rg_date')
            elif sort == 'asaction':
                company_sheet = ASsheet.objects.filter(cname=login_session).order_by('-asaction', '-rg_date')
            elif sort == 'la_category':
                company_sheet = ASsheet.objects.filter(cname=login_session).order_by('-la_category', '-rg_date')
            elif sort == 'me_category':
                company_sheet = ASsheet.objects.filter(cname=login_session).order_by('-me_category', '-rg_date')
            elif sort == 'sm_category':
                company_sheet = ASsheet.objects.filter(cname=login_session).order_by('-sm_category', '-rg_date')
            elif sort == 'user_name':
                company_sheet = ASsheet.objects.filter(cname=login_session).order_by('-user_name', '-rg_date')
            elif sort == 'all':
                company_sheet = ASsheet.objects.filter(cname=login_session).order_by('-rg_date', 'finish')
            else:
                print("리스트 조회 겸 목록 조회")
                search_sort = request.GET.get('search_sort', '')
                startdate = request.GET.get('sdate', '')
                enddate = request.GET.get('edate', '')
                if search_sort == 'cname':
                    company_sheet = ASsheet.objects.all().filter(cname__icontains=query, cname=login_session).order_by(
                        '-rg_date',
                        'finish')
                elif search_sort == 'cuser':
                    company_sheet = ASsheet.objects.all().filter(cuser__icontains=query, cname=login_session).order_by(
                        '-rg_date',
                        'finish')
                elif search_sort == 'product_name':
                    company_sheet = ASsheet.objects.all().filter(product_name__icontains=query,
                                                                 cname=login_session).order_by('-rg_date',
                                                                                               'finish')
                elif search_sort == 'serial':
                    company_sheet = ASsheet.objects.all().filter(serial__icontains=query, cname=login_session).order_by(
                        '-rg_date',
                        'finish')
                elif search_sort == 'after_serial':
                    company_sheet = ASsheet.objects.all().filter(after_serial__icontains=query,
                                                                 cname=login_session).order_by('-rg_date',
                                                                                               'finish')
                elif search_sort == 'asaction':
                    company_sheet = ASsheet.objects.all().filter(asaction__icontains=query,
                                                                 cname=login_session).order_by('-rg_date',
                                                                                               'finish')
                elif search_sort == 'la_category':
                    company_sheet = ASsheet.objects.all().filter(la_category__icontains=query,
                                                                 cname=login_session).order_by('-rg_date',
                                                                                               'finish')
                elif search_sort == 'me_category':
                    company_sheet = ASsheet.objects.all().filter(me_category__icontains=query,
                                                                 cname=login_session).order_by('-rg_date',
                                                                                               'finish')
                elif search_sort == 'sm_category':
                    company_sheet = ASsheet.objects.all().filter(sm_category__icontains=query,
                                                                 cname=login_session).order_by('-rg_date',
                                                                                               'finish')
                elif search_sort == 'all':
                    company_sheet = ASsheet.objects.all().filter(
                        Q(product_name__icontains=query) | Q(memo__icontains=query) | Q(cname__icontains=query)
                        | Q(finish__icontains=query) | Q(cuser__icontains=query) | Q(cphone__icontains=query)
                        | Q(option__icontains=query) | Q(serial__icontains=query) | Q(
                            symptom__icontains=query), cname=login_session).order_by('-rg_date', 'finish')
                elif search_sort == 'rg_date':
                    e_date = datetime.datetime.strptime(enddate, '%Y-%m-%d') + datetime.timedelta(hours=23, minutes=59,
                                                                                                  seconds=59)
                    company_sheet = ASsheet.objects.all().filter(rg_date__gte=startdate, rg_date__lte=e_date,
                                                                 cname=login_session).order_by(
                        '-rg_date', 'finish')
                elif search_sort == 'rp_date':
                    company_sheet = ASsheet.objects.all().filter(rp_date__range=[startdate, enddate],
                                                                 cname=login_session).order_by(
                        '-rg_date', 'finish')
                elif search_sort == 'end_date':
                    company_sheet = ASsheet.objects.all().filter(end_date__range=[startdate, enddate],
                                                                 cname=login_session).order_by(
                        '-rg_date', 'finish')
                else:
                    company_sheet = ASsheet.objects.all().filter(cname=login_session).order_by('-rg_date', 'finish')

            page = request.GET.get('page', '1')
            paginator = Paginator(company_sheet, 10)
            page_obj = paginator.get_page(page)
            print("GET 페이징 끝")
            context = {'login_session': login_session, 'company_sheet': company_sheet, 'page_obj': page_obj,
                       'user_phone': user_phone,
                       'sort': sort, 'search_sort': search_sort, 'query': query, 'user_name': user_name,
                       'sdate': startdate, 'edate': enddate}
        print('끝')
        return render(request, 'assheet/as_list.html', context)
    elif request.method == 'POST':
        print('포스트인가')
        print("포스트 리스트 끝")
        return redirect('asregister:as_list')


# AS 파일 업로드/다운로드
def AsUploadFile(request, pk):
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')
    if request.method == "POST":
        if request.FILES.get('uploadedFile') is not None:
            if get_object_or_404(ASsheet, no=pk):
                # 템플릿에서 데이터 가져오기
                cname = login_session
                sheet_no = ASsheet.objects.get(no=pk)
                menu = request.POST["menu"]

                files = request.FILES.getlist('uploadedFile')

                for f in files:
                    fileTitle = f
                    uploadedFile = f
                    # DB에 저장
                    uploadfile = models.ASUploadFile(
                        cname=cname,
                        title=fileTitle,
                        uploadedFile=uploadedFile,
                        sheet_no=sheet_no,
                        menu=menu
                    )
                    uploadfile.save()
                return HttpResponseRedirect(reverse('asregister:AsUploadFile', args=[pk]))
    else:
        login_session = request.session.get('login_session')
        detailView = get_object_or_404(ASsheet, no=pk)
        uploadfile = models.ASUploadFile.objects.all()
        no = pk

        context = {'login_session': login_session, 'no': no, 'detailView': detailView, 'files': uploadfile,
                   'user_name': user_name}
        return render(request, "assheet/asfile_upload.html", context)

    uploadfile = models.ASUploadFile.objects.all()
    detailView = get_object_or_404(ASsheet, no=pk)

    return render(request, "assheet/asfile_upload.html", context={'user_name': user_name,
                                                                  "files": uploadfile, "login_session": login_session,
                                                                  'detailView': detailView})


# as파일 다운로드
def as_downloadfile(request, pk):
    upload_file = get_object_or_404(ASUploadFile, no=pk)
    file = upload_file.uploadedFile
    name = file.name
    response = HttpResponse(content_type=mimetypes.guess_type(name)[0] or 'application/octet-stream')
    response['Content-Disposition'] = f'attachment; filename={name}'
    shutil.copyfileobj(file, response)
    return response


# AS 액셀 다운로드 openpyxl 사용
def AS_excel_openpyxl(request):
    login_session = request.session.get('login_session')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response["Content-Disposition"] = 'attachment; filename=' \
                                      + str(datetime.date.today()) + '_AS.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AS'

    columns = ['업체명', '요청자', '연락처', '접수 일자', '마감 요청 일자', '종료 일자', '제품명',
               '전 시리얼', '후 시리얼', '프로젝트명', '증상', '비고', '의견', '품목', '대분류', '소분류', '조치', '담당자', '완료 여부']

    query = request.GET.get('q')
    search_sort = request.GET.get('search_sort', '')
    if search_sort:
        startdate = request.GET.get('sdate', '')
        enddate = request.GET.get('edate', '')
        if login_session == 'insung':
            if search_sort == 'product_name':
                rows = ASsheet.objects.all().filter(product_name__icontains=query)
            elif search_sort == 'cname':
                rows = ASsheet.objects.all().filter(cname__icontains=query)
            elif search_sort == 'cuser':
                rows = ASsheet.objects.all().filter(cuser__icontains=query)
            elif search_sort == 'cphone':
                rows = ASsheet.objects.all().filter(cphone__icontains=query)
            elif search_sort == 'finish':
                rows = ASsheet.objects.all().filter(finish__icontains=query)
            elif search_sort == 'memo':
                rows = ASsheet.objects.all().filter(memo__icontains=query)
            elif search_sort == 'serial':
                rows = ASsheet.objects.all().filter(serial__icontains=query)
            elif search_sort == 'after_serial':
                rows = ASsheet.objects.all().filter(after_serial__icontains=query)
            elif search_sort == 'all':
                rows = ASsheet.objects.all().filter(Q(product_name__icontains=query) | Q(cname__icontains=query) |
                                                    Q(cuser__icontains=query) | Q(cphone__icontains=query) |
                                                    Q(finish__icontains=query) | Q(memo__icontains=query) |
                                                    Q(serial__icontains=query) | Q(site__icontains=query) |
                                                    Q(symptom__icontains=query) | Q(option__icontains=query) |
                                                    Q(option__icontains=query))
            elif search_sort == 'site':
                rows = ASsheet.objects.all().filter(site__icontains=query)
            elif search_sort == 'symptom':
                rows = ASsheet.objects.all().filter(symptom__icontains=query)
            elif search_sort == 'option':
                rows = ASsheet.objects.all().filter(option__icontains=query)
            elif search_sort == 'rg_date':
                e_date = datetime.datetime.strptime(enddate, '%Y-%m-%d') + datetime.timedelta(hours=23, minutes=59,
                                                                                              seconds=59)
                rows = ASsheet.objects.all().filter(rg_date__gte=startdate, rg_date__lte=e_date).order_by(
                    '-rg_date', 'finish')
            elif search_sort == 'rp_date':
                rows = ASsheet.objects.all().filter(rp_date__range=[startdate, enddate]).order_by('-rg_date', 'finish')
            elif search_sort == 'end_date':
                rows = ASsheet.objects.all().filter(end_date__range=[startdate, enddate]).order_by('-rg_date',
                                                                                                   'finish')
            else:
                rows = ASsheet.objects.all()
        else:
            if search_sort == 'product_name':
                rows = ASsheet.objects.all().filter(product_name__icontains=query, cname=login_session)
            elif search_sort == 'cname':
                rows = ASsheet.objects.all().filter(cname__icontains=query, cname=login_session)
            elif search_sort == 'cuser':
                rows = ASsheet.objects.all().filter(cuser__icontains=query, cname=login_session)
            elif search_sort == 'cphone':
                rows = ASsheet.objects.all().filter(cphone__icontains=query, cname=login_session)
            elif search_sort == 'finish':
                rows = ASsheet.objects.all().filter(finish__icontains=query, cname=login_session)
            elif search_sort == 'memo':
                rows = ASsheet.objects.all().filter(memo__icontains=query, cname=login_session)
            elif search_sort == 'serial':
                rows = ASsheet.objects.all().filter(serial__icontains=query, cname=login_session)
            elif search_sort == 'after_serial':
                rows = ASsheet.objects.all().filter(after_serial__icontains=query, cname=login_session)
            elif search_sort == 'all':
                rows = ASsheet.objects.all().filter(Q(product_name__icontains=query) | Q(cname__icontains=query) |
                                                    Q(cuser__icontains=query) | Q(cphone__icontains=query) |
                                                    Q(finish__icontains=query) | Q(memo__icontains=query) |
                                                    Q(serial__icontains=query) | Q(site__icontains=query) |
                                                    Q(symptom__icontains=query) | Q(option__icontains=query) |
                                                    Q(option__icontains=query), cname=login_session)
            elif search_sort == 'site':
                rows = ASsheet.objects.all().filter(site__icontains=query, cname=login_session)
            elif search_sort == 'symptom':
                rows = ASsheet.objects.all().filter(symptom__icontains=query, cname=login_session)
            elif search_sort == 'option':
                rows = ASsheet.objects.all().filter(option__icontains=query, cname=login_session)
            elif search_sort == 'rg_date':
                e_date = datetime.datetime.strptime(enddate, '%Y-%m-%d') + datetime.timedelta(hours=23, minutes=59,
                                                                                              seconds=59)
                rows = ASsheet.objects.all().filter(rg_date__gte=startdate, rg_date__lte=e_date,
                                                    cname=login_session).order_by(
                    '-rg_date', 'finish')
            elif search_sort == 'rp_date':
                rows = ASsheet.objects.all().filter(rp_date__range=[startdate, enddate], cname=login_session).order_by(
                    '-rg_date', 'finish')
            elif search_sort == 'end_date':
                rows = ASsheet.objects.all().filter(end_date__range=[startdate, enddate], cname=login_session).order_by(
                    '-rg_date', 'finish')
            else:
                rows = ASsheet.objects.all().filter(cname=login_session)
    else:
        if login_session == 'insung':
            # 엑셀에 쓸 데이터 리스트화
            rows = ASsheet.objects.all()
        else:
            rows = ASsheet.objects.filter(cname=login_session)

    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title

    for asrow in rows:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            asrow.cname,
            asrow.cuser,
            asrow.cphone,
            asrow.rg_date.strftime('%Y-%m-%d'),
            asrow.rp_date,
            asrow.end_date,
            asrow.product_name,
            asrow.serial,
            asrow.site,
            asrow.after_serial,
            asrow.symptom,
            asrow.memo,
            asrow.option,
            asrow.la_category,
            asrow.me_category,
            asrow.sm_category,
            asrow.asaction,
            asrow.finish,
            asrow.user_name
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value

    wb.save(response)
    return response


# as 파일 삭제
def ASfile_delete(request, pk):
    login_session = request.session.get('login_session')
    if request.method == 'GET':
        print("get 으로 옴")
        as_file = get_object_or_404(ASUploadFile, no=pk)
        page_no = as_file.sheet_no_id
        if as_file.cname == login_session or login_session == 'insung':
            as_file.delete()
            print('삭제완료')
            return redirect(f'/asregister/AsUploadFile/{page_no}')
        else:
            print("삭제 됨?")
            return redirect('/asregister/as_list/')
    else:
        print('post로 옴')


# AS 접수서 상세 뷰
def as_detail(request, pk):
    if request.method == 'GET':
        login_session = request.session.get('login_session')
        user_name = request.session.get('user_name')
        user_phone = request.session.get('user_phone')
        detailView = get_object_or_404(ASsheet, no=pk)

        sort = request.GET.get('sort', '')
        query = request.GET.get('q', '')
        search_sort = request.GET.get('search_sort', '')
        if request.GET.get('sdate', '') is not None:
            startdate = request.GET.get('sdate', '')
        if request.GET.get('edate', '') is not None:
            enddate = request.GET.get('edate', '')
        page = request.GET.get('page', '')

        try:
            upfile = ASUploadFile.objects.filter(sheet_no_id=pk).count()
            context = {'detailView': detailView, 'login_session': login_session, 'upfile': upfile,
                       'user_name': user_name, 'user_phone': user_phone, 'sort': sort, 'query': query,
                       'search_sort': search_sort,
                       'sdate': startdate, 'edate': enddate, 'page': page}
            print("성공")
        except:
            context = {'detailView': detailView, 'login_session': login_session, 'user_name': user_name,
                       'user_phone': user_phone, 'sort': sort, 'query': query, 'search_sort': search_sort,
                       'sdate': startdate, 'edate': enddate, 'page': page}
            print("실패")

    else:
        print("설마 포스트")
    return render(request, 'assheet/as_detail.html', context)


# AS 접수건 응대 / 업데이트
def as_modify(request, pk):
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')
    user_phone = request.session.get('user_phone')
    # render context로 넘길때 key:value 로 넘겨야 넘어가고 받아진다

    detailView = get_object_or_404(ASsheet, no=pk)

    if request.method == 'GET':
        # get으로 오면 다시 수정페이지로 넘김
        detailView = get_object_or_404(ASsheet, no=pk)

        context = {'detailView': detailView, 'login_session': login_session, 'user_name': user_name,
                   'user_phone': user_phone}
        return render(request, 'assheet/as_modify.html', context)
    elif request.method == 'POST':

        # 수정 내용 저장
        detailView.rg_date = request.POST['rg_date']
        detailView.product_name = request.POST['product_name']
        detailView.cname = request.POST['cname']
        detailView.serial = request.POST['serial']
        detailView.user_name = user_name
        detailView.memo = request.POST['memo']
        detailView.option = request.POST['option']
        detailView.finish = request.POST['finish']
        if login_session == 'insung':
            print('입력')
            detailView.after_serial = request.POST.get('after_serial', None)
            detailView.la_category = request.POST.get('la', None)
            detailView.me_category = request.POST.get('me', None)
            detailView.sm_category = request.POST.get('sm', None)
            detailView.asaction = request.POST.get('action', None)
        if request.POST['finish'] == '종료':
            if request.POST['after_serial'] == '':
                print('확인')
                detailView.end_date = date.today()
                try:
                    print('pm 수정 시작')
                    ex_pm = Product_Management.objects.exclude(status='폐기')
                    pm_modify = get_object_or_404(ex_pm, serial=request.POST['serial'])
                    pm_modify2 = get_object_or_404(ex_pm, serial=request.POST['after_serial'])
                    pm_modify.current_location = request.POST['cname']
                    pm_modify2.current_location = request.POST['cname']
                    pm_modify.status = "AS"
                    pm_modify2.status = "출고"

                    pm_modify.save()
                    pm_modify2.save()
                    print("pm 까지 수정 저장완료")
                except:
                    print('후 시리얼 없어 예외처리')
                    detailView.save()
            elif request.POST['serial'] == '':
                print('후 시리얼 없어 전 시리얼만 업데이트')
                ex_pm = Product_Management.objects.exclude(status='폐기')
                pm_modify = get_object_or_404(ex_pm, serial=request.POST['serial'])
                pm_modify.current_location = request.POST['cname']
                pm_modify.status = "AS"
                pm_modify.save()
            else:
                print('전 시리얼 없이 저장')
                detailView.save()
        elif request.POST['finish'] == '진행 중' or request.POST['finish'] == '접수 중':
            try:
                ex_pm = Product_Management.objects.exclude(status='폐기')
                pm_modify = get_object_or_404(ex_pm, serial=request.POST['serial'])
                pm_modify.current_location = request.POST['cname']
                pm_modify.status = "AS"

                pm_modify.save()
                print("pm 까지 수정 저장완료")
            except:
                print("수정 저장완료")
                detailView.save()

        detailView.save()
        print('끝')
        detailView = get_object_or_404(ASsheet, no=pk)
        context = {'detailView': detailView, 'login_session': login_session, 'user_name': user_name,
                   'user_phone': user_phone}
        return render(request, 'assheet/as_detail.html', context)


# AS 접수건 삭제
def as_delete(request, pk):
    login_session = request.session.get('login_session')
    detailView = get_object_or_404(ASsheet, no=pk)
    if detailView.cname == login_session or login_session == 'insung':
        detailView.delete()
        print('삭제완료')
        return redirect('asregister:as_list')
    else:
        return redirect(f'/asregister/as_detail/{pk}')


# AS 리포트
def as_report(request):
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')

    # 주간 월간 제품별 AS 현황
    as_wnum = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(weeks=1)).values('product_name'). \
        order_by('product_name').annotate(count=Count('product_name'))
    as_wnum_sum = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(weeks=1)).aggregate(
        Count('product_name'))
    as_mnum = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1)).values('product_name'). \
        order_by('product_name').annotate(count=Count('product_name'))
    as_mnum_sum = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1)).aggregate(
        Count('product_name'))

    # 주간 월간 품목/대,소/조치 AS 현황
    # 주간
    as_wle = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(weeks=1)).exclude(
        la_category=None).values('la_category'). \
        order_by('la_category').annotate(count=Count('la_category'))
    as_wle_sum = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(weeks=1)).exclude(
        la_category=None).aggregate(Count('la_category'))

    as_wme = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(weeks=1)).exclude(
        me_category=None).values('me_category'). \
        order_by('me_category').annotate(count=Count('me_category'))
    as_wme_sum = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(weeks=1)).exclude(
        me_category=None).aggregate(
        Count('me_category'))

    as_wsm = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(weeks=1)).exclude(
        sm_category=None).values('sm_category'). \
        order_by('sm_category').annotate(count=Count('sm_category'))
    as_wsm_sum = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(weeks=1)).exclude(
        sm_category=None).aggregate(
        Count('sm_category'))

    as_wac = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(weeks=1)).exclude(asaction=None).values(
        'asaction'). \
        order_by('asaction').annotate(count=Count('asaction'))
    as_wac_sum = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(weeks=1)).exclude(
        asaction=None).aggregate(
        Count('asaction'))

    # 월간
    as_mle = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1)).exclude(
        la_category=None).values('la_category'). \
        order_by('la_category').annotate(count=Count('la_category'))
    as_mle_sum = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1)).exclude(
        la_category=None).aggregate(
        Count('la_category'))

    as_mme = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1)).exclude(
        me_category=None).values('me_category'). \
        order_by('me_category').annotate(count=Count('me_category'))
    as_mme_sum = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1)).exclude(
        me_category=None).aggregate(
        Count('me_category'))

    as_msm = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1)).exclude(
        sm_category=None).values('sm_category'). \
        order_by('sm_category').annotate(count=Count('sm_category'))
    as_msm_sum = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1)).exclude(
        sm_category=None).aggregate(
        Count('sm_category'))

    as_mac = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1)).exclude(asaction=None).values(
        'asaction'). \
        order_by('asaction').annotate(count=Count('asaction'))
    as_mac_sum = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1)).exclude(
        asaction=None).aggregate(
        Count('asaction'))

    context = {'as_wnum': as_wnum, 'as_wnum_sum': as_wnum_sum, 'as_mnum': as_mnum, 'as_mnum_sum': as_mnum_sum,
               'as_wle': as_wle,
               'as_wle_sum': as_wle_sum, 'as_wme': as_wme, 'as_wme_sum': as_wme_sum, 'as_wsm': as_wsm,
               'as_wsm_sum': as_wsm_sum,
               'as_wac': as_wac, 'as_wac_sum': as_wac_sum, 'as_mle': as_mle, 'as_mle_sum': as_mle_sum, 'as_mme': as_mme,
               'as_mme_sum': as_mme_sum, 'as_msm': as_msm, 'as_msm_sum': as_msm_sum, 'as_mac': as_mac,
               'as_mac_sum': as_mac_sum,
               'login_session': login_session, 'user_name': user_name}

    return render(request, 'assheet/as_report.html', context)


# AS 차트 페이지
def as_chart(request):
    global chart
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')

    search_sort = request.GET.get('search_sort', '')
    startdate = request.GET.get('sdate', '')
    enddate = request.GET.get('edate', '')

    print('search : ', search_sort)
    print('sdate : ', startdate)
    print('edate : ', enddate)

    la_category = request.GET.get('la', '')
    print('la : ', la_category)
    me_category = request.GET.get('me', '')
    print('me : ', me_category)
    sm_category = request.GET.get('sm', '')
    print('sm : ', sm_category)
    asaction = request.GET.get('action', '')
    print('as : ', asaction)
    print('여기?')
    if search_sort == 'rg_date':
        if la_category == '' and me_category == '' and sm_category == '' and asaction == '':
            print('여기')
            e_date = datetime.datetime.strptime(enddate, '%Y-%m-%d') + datetime.timedelta(hours=23, minutes=59,
                                                                                          seconds=59)
            chart = ASsheet.objects.all().filter(rg_date__gte=startdate, rg_date__lte=e_date).order_by('-rg_date',
                                                                                                       'finish')
        else:
            print('여기2')
            e_date = datetime.datetime.strptime(enddate, '%Y-%m-%d') + datetime.timedelta(hours=23, minutes=59,
                                                                                          seconds=59)
            chartdata = ASsheet.objects.all().filter(rg_date__gte=startdate, rg_date__lte=e_date).order_by('-rg_date',
                                                                                                           'finish')
            chart = chartdata.filter(Q(la_category__exact=la_category) | Q(me_category__exact=me_category) | \
                                     Q(sm_category__exact=sm_category) | Q(asaction__exact=asaction)).order_by('-rg_date', 'finish')
            print('여기 끝')
    elif search_sort == 'rp_date':
        if la_category == '' and me_category == '' and sm_category == '' and asaction == '':
            chart = ASsheet.objects.all().filter(rp_date__range=[startdate, enddate]).order_by('-rg_date', 'finish')
        else:
            chartdata = ASsheet.objects.all().filter(rp_date__range=[startdate, enddate]).order_by('-rg_date', 'finish')
            chart = chartdata.filter(
                Q(la_category__exact=la_category) | Q(me_category__exact=me_category) |
                Q(sm_category__exact=sm_category) | Q(asaction__exact=asaction)).order_by('-rg_date', 'finish')
    elif search_sort == 'end_date':
        if la_category == '' and me_category == '' and sm_category == '' and asaction == '':
            chart = ASsheet.objects.all().filter(end_date__range=[startdate, enddate]).order_by('-rg_date', 'finish')
        else:
            chartdata = ASsheet.objects.all().filter(end_date__range=[startdate, enddate]).order_by('-rg_date',
                                                                                                    'finish')
            chart = chartdata.filter(
                Q(la_category__exact=la_category) | Q(me_category__exact=me_category) |
                Q(sm_category__exact=sm_category) | Q(asaction__exact=asaction)).order_by('-rg_date', 'finish')

    sheet_chart = []
    context = {'login_session': login_session, 'user_name': user_name, 'chart': chart.count()}

    return render(request, 'assheet/as_report.html', context)
