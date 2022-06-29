from django.shortcuts import render, redirect, get_object_or_404, reverse
from .decorators import login_required, login_ok
import sys

sys.path.append('..')
# 빨간줄 나타나지만 정상작동 / 표현만 저렇게 되는 것
from asregister.models import ASsheet
from question.models import question_sheet
from .models import UploadFile, ProductDb, main_sheet, sub_sheet, product_info, notice, Product_Management
from . import models
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count
from django.http import JsonResponse, HttpResponseRedirect
import datetime
import xlwt
import openpyxl
from django.http import HttpResponse
import mimetypes
import shutil
from datetime import date
from dateutil.relativedelta import relativedelta
from functools import reduce
from django.urls import reverse
from openpyxl import load_workbook


# Create your views here.

# 메인페이지
def index(request):
    login_session = request.session.get('login_session')
    user_phone = request.session.get('user_phone')
    user_name = request.session.get('user_name')

    # 당일 기준 신규 접수 현황
    es_count = main_sheet.objects.filter(rg_date__gte=date.today()).count()
    es_pcount = main_sheet.objects.filter(rg_date__gte=date.today(),
                                          cname=login_session).count()
    es_fcount = main_sheet.objects.filter(end_date__gte=date.today(),
                                          finish="종료").count()
    es_pfcount = main_sheet.objects.filter(end_date__gte=date.today(),
                                           finish="종료", cname=login_session).count()
    es_icount = main_sheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1),
                                          finish="진행 중").count()
    es_picount = main_sheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1),
                                           finish="진행 중", cname=login_session).count()
    es_xcount = main_sheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1),
                                          finish="접수 중").count()
    es_pxcount = main_sheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1),
                                           finish="접수 중", cname=login_session).count()
    as_count = ASsheet.objects.filter(rg_date__gte=date.today()).count()
    as_pcount = ASsheet.objects.filter(rg_date__gte=date.today(),
                                       cname=login_session).count()
    as_fcount = ASsheet.objects.filter(end_date__gte=date.today(), finish="종료").count()
    as_pfcount = ASsheet.objects.filter(end_date__gte=date.today(), finish="종료", cname=login_session).count()
    as_icount = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1), finish="진행 중").count()
    as_picount = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1), finish="진행 중",
                                        cname=login_session).count()
    as_xcount = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1), finish="접수 중").count()
    as_pxcount = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1), finish="접수 중",
                                        cname=login_session).count()
    que_count = question_sheet.objects.filter(rg_date__gte=date.today()).count()
    que_pcount = question_sheet.objects.filter(rg_date__gte=date.today(),
                                               cname=login_session).count()

    # 주간 실적 현황
    es_week1 = main_sheet.objects.filter(rg_date__gte=date.today() - datetime.timedelta(weeks=1),
                                         user_dept='영업1팀').aggregate(Sum('total_price'))
    es_week2 = main_sheet.objects.filter(rg_date__gte=date.today() - datetime.timedelta(weeks=1),
                                         user_dept='영업2팀').aggregate(Sum('total_price'))

    # 월간 실적 현황
    es_month1 = main_sheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1),
                                          user_dept='영업1팀').aggregate(Sum('total_price'))
    es_month2 = main_sheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1),
                                          user_dept='영업2팀').aggregate(Sum('total_price'))

    # 제품별 월간 제품 출고, AS 개수
    es_num = sub_sheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1)).values(
        'product_name').order_by('product_name').annotate(count=Sum('quantity'))
    es_num_sum = sub_sheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1)).aggregate(
        Sum('quantity'))
    as_num = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1)).values(
        'product_name').order_by('product_name').annotate(count=Count('product_name'))
    as_num_sum = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1)).aggregate(
        Count('product_name'))

    # 업체별 월간 메인, AS 개수
    es_cnum = main_sheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1)).values(
        'cname').order_by('cname').annotate(count=Count('cname'))
    es_cnum_sum = main_sheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1)).values(
        'cname').order_by('cname').annotate(count=Count('cname')).aggregate(Sum('count'))
    as_cnum = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1)).values('cname').order_by(
        'cname').annotate(count=Count('cname'))
    as_cnum_sum = ASsheet.objects.filter(rg_date__gte=date.today() - relativedelta(months=1)).values(
        'cname').order_by('cname').annotate(count=Count('cname')).aggregate(Sum('count'))

    # 공지사항 리스트
    notice_list = notice.objects.all().order_by('-rg_date')
    page = request.GET.get('page', '1')
    paginator = Paginator(notice_list, 5)
    page_obj = paginator.get_page(page)

    context = {'login_session': login_session, 'es_count': es_count, 'as_count': as_count,
               'que_count': que_count, 'es_icount': es_icount,
               'es_week1': es_week1, 'es_week2': es_week2,
               'es_month1': es_month1, 'es_month2': es_month2,
               'es_num': es_num, 'as_num': as_num, 'es_xcount': es_xcount,
               'es_cnum': es_cnum, 'as_cnum': as_cnum, 'es_fcount': es_fcount, 'as_xcount': as_xcount,
               'as_icount': as_icount,
               'as_fcount': as_fcount, 'es_pcount': es_pcount,
               'as_pcount': as_pcount, 'que_pcount': que_pcount, 'es_num_sum': es_num_sum,
               'as_num_sum': as_num_sum, 'es_cnum_sum': es_cnum_sum,
               'as_cnum_sum': as_cnum_sum, 'es_pfcount': es_pfcount, 'es_picount': es_picount, 'es_pxcount': es_pxcount,
               'as_pfcount': as_pfcount, 'as_picount': as_picount, 'as_pxcount': as_pxcount, 'user_name': user_name,
               'page_obj': page_obj}
    return render(request, 'isscm/index.html', context)


# 제품명 검색 자동완성
def searchData(request):
    if 'term' in request.GET:
        qs = ProductDb.objects.filter(product_name__icontains=request.GET.get('term'))
        pname = list()
        if qs:
            for product in qs:
                pname.append(product.product_name)
                print(pname)
        else:
            print("없음")
            n = '일치하는 제품명이 없습니다.'
            pname.append(n)
        return JsonResponse(pname, safe=False)


# 제품명 검색 자동완성(시리얼로 제품명 검색, 중복 제거)
def searchPM(request):
    if 'term' in request.GET:
        qs = Product_Management.objects.filter(serial__icontains=request.GET.get('term')).exclude(status='폐기').values(
            'product_name').distinct()
        pname = list()
        pmlist = []
        print(qs.values())
        if qs:
            for product in qs:
                n = product['product_name']
                pname.append(n)
        else:
            print("없음")
            n = '시리얼과 일치하는 제품명이 없습니다.'
            pname.append(n)
        return JsonResponse(pname, safe=False)


# 시리얼 검색(중복제거)
def searchPM_serial(request):
    if 'term' in request.GET:
        qs = Product_Management.objects.filter(serial__icontains=request.GET.get('term')).exclude(status='폐기').values(
            'serial').distinct()
        p1 = list()
        pmlist = []
        print("여기는? ", qs.values())
        if qs:
            for product in qs:
                n = product['serial']
                p1.append(n)
        else:
            print("없음")
            n = '일치하는 시리얼이 없습니다.'
            p1.append(n)
        return JsonResponse(p1, safe=False)


# 메인 입력
@login_required
def main_insert(request):
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')
    user_dept = request.session.get('user_dept')

    if request.method == 'GET':
        print('메인 입력 겟 들어옴')
        context = {'login_session': login_session, 'user_name': user_name, 'user_dept': user_dept}
        print('메인 입력 겟 나감')
        return render(request, 'isscm/main_insert.html', context)
    elif request.method == 'POST':
        print("메인 입력 시작")
        main = main_sheet()
        main.rp_date = request.POST['rp_date']
        main.main_title = request.POST['main_title']
        main.cname = request.POST['cname']
        main.requests = request.POST['requests']
        main.total_price = request.POST.get('total_price')
        main.finish = request.POST.get('finish', '')
        main.user_name = request.POST.get('user_name', '')
        main.user_dept = request.POST.get('user_dept', '')

        main.save()

        # 값 넘길 필요없어 미사용
        context = {'login_session': login_session, 'user_name': user_name, 'user_dept': user_dept}
        print('메인 입력 끝남')
        return redirect('isscm:main_list')


def main_detail(request, pk):
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')
    user_dept = request.session.get('user_dept')
    detailView = get_object_or_404(main_sheet, id=pk)
    sort = request.GET.get('sort', '')
    query = request.GET.get('q', '')
    search_sort = request.GET.get('search_sort', '')
    if request.GET.get('sdate', '') is not None:
        startdate = request.GET.get('sdate', '')
    if request.GET.get('edate', '') is not None:
        enddate = request.GET.get('edate', '')
    page = request.GET.get('page', '')

    global sub
    if request.method == 'GET':
        print('get 메인 디테일 뷰 시작')
        if sub_sheet.objects.filter(m_id_id=pk):

            sub = sub_sheet.objects.filter(m_id_id=pk).order_by('-rg_date')
            # 목록으로 돌아갈때 페이지 정보 필요하여 넘김
            page = request.GET.get('page', '1')
            print("insung GET main 페이징 끝")
            try:
                # 업로드 파일 있을때 파일 갯수 넘김
                upload_file = UploadFile.objects.filter(main_id_id=pk).count()
                context = {'detailView': detailView, 'login_session': login_session, 'user_name': user_name,
                           'user_dept': user_dept, 'sub': sub, 'files': upload_file, 'sort': sort, 'query': query,
                           'search_sort': search_sort,
                           'sdate': startdate, 'edate': enddate, 'page': page}
            except:
                print('get 파일 없음')
                context = {'detailView': detailView, 'login_session': login_session, 'user_name': user_name,
                           'user_dept': user_dept, 'sub': sub, 'sort': sort, 'query': query, 'search_sort': search_sort,
                           'sdate': startdate, 'edate': enddate, 'page': page}
        else:
            print('sub 없음')
            upload_file = UploadFile.objects.filter(main_id_id=pk).count()
            context = {'detailView': detailView, 'login_session': login_session, 'user_name': user_name,
                       'user_dept': user_dept, 'files': upload_file, 'sort': sort, 'query': query,
                       'search_sort': search_sort,
                       'sdate': startdate, 'edate': enddate, 'page': page}
        print("디테일 뷰 끝")
        return render(request, 'isscm/main_detail.html', context)
    else:
        print("post 메인 디테일 뷰 / 수정 시작")
        # 수정 내용 저장
        detailView.rp_date = request.POST['rp_date']
        detailView.main_title = request.POST['main_title']
        detailView.total_price = request.POST.get('total_price').replace(",", "")
        detailView.cname = request.POST['cname']
        detailView.requests = request.POST['requests']
        detailView.finish = request.POST.get('finish', None)
        detailView.user_dept = request.POST.get('user_dept')
        detailView.user_name = request.POST.get('user_name')
        if request.POST.get('finish') == '종료':
            detailView.end_date = date.today()
        detailView.save()

        return HttpResponseRedirect(reverse('isscm:main_list'))


# main 삭제
def main_delete(request, pk):
    detailView = get_object_or_404(main_sheet, id=pk)
    detailView.delete()
    print('삭제완료')
    return redirect('isscm:main_list')


# main 리스트
@login_required
def main_list(request):
    print("main 리스트 시작")
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')
    user_dept = request.session.get('user_dept')
    global search_sort
    global startdate
    global enddate
    if request.method == 'GET':
        if login_session == 'insung':
            print('get insung 리스트 시작')
            sort = request.GET.get('sort', '')
            query = request.GET.get('q', '')
            if sort == 'rg_date':
                m_sheet = main_sheet.objects.all().order_by('-rg_date')
            elif sort == 'rp_date':
                m_sheet = main_sheet.objects.all().order_by('-rp_date')
            elif sort == 'main_title':
                m_sheet = main_sheet.objects.all().order_by('-main_title', '-rg_date')
            elif sort == 'cname':
                m_sheet = main_sheet.objects.all().order_by('-cname', '-rg_date', '-finish')
            elif sort == 'finish':
                m_sheet = main_sheet.objects.all().order_by('-finish', '-rg_date')
            elif sort == 'total_price':
                m_sheet = main_sheet.objects.all().order_by('-total_price', '-rg_date')
            elif sort == 'user_dept':
                m_sheet = main_sheet.objects.all().order_by('-user_dept', '-rg_date')
            elif sort == 'user_name':
                m_sheet = main_sheet.objects.all().order_by('-user_name', '-rg_date')
            elif sort == 'all':
                m_sheet = main_sheet.objects.all().order_by('-rg_date', 'finish', '-user_dept')
            else:
                print("리스트 조회 겸 목록 조회")
                search_sort = request.GET.get('search_sort', '')
                startdate = request.GET.get('sdate', '')
                enddate = request.GET.get('edate', '')
                if search_sort == 'main_title':
                    m_sheet = main_sheet.objects.all().filter(main_title__icontains=query).order_by('-rg_date')
                elif search_sort == 'requests':
                    m_sheet = main_sheet.objects.all().filter(requests__icontains=query).order_by('-rg_date')
                elif search_sort == 'cname':
                    m_sheet = main_sheet.objects.all().filter(cname__icontains=query).order_by('-rg_date')
                elif search_sort == 'user_dept':
                    m_sheet = main_sheet.objects.all().filter(user_dept__icontains=query).order_by('-rg_date')
                elif search_sort == 'user_name':
                    m_sheet = main_sheet.objects.all().filter(user_name__icontains=query).order_by('-rg_date')
                elif search_sort == 'product_name':
                    s_sheet = sub_sheet.objects.all().filter(product_name__icontains=query).order_by('-rg_date')
                    v = []
                    for i in s_sheet:
                        sm = i.m_id_id
                        print(sm)
                        ms = main_sheet.objects.filter(id__iexact=sm)
                        v = v + list(ms)
                    # 중복제거 / 정렬 작업을 위한 조건문
                    m_sheet = reduce(lambda acc, cur: acc if cur in acc else acc + [cur], v, [])
                elif search_sort == 'rg_date':
                    e_date = datetime.datetime.strptime(enddate, '%Y-%m-%d') + datetime.timedelta(hours=23, minutes=59,
                                                                                                  seconds=59)
                    m_sheet = main_sheet.objects.all().filter(rg_date__gte=startdate, rg_date__lte=e_date).order_by(
                        '-rg_date')
                elif search_sort == 'rp_date':
                    m_sheet = main_sheet.objects.all().filter(rp_date__range=[startdate, enddate]).order_by('-rg_date')
                elif search_sort == 'end_date':
                    m_sheet = main_sheet.objects.all().filter(end_date__range=[startdate, enddate]).order_by('-rg_date')
                elif search_sort == 'all':
                    m_sheet = main_sheet.objects.filter(
                        Q(main_title__icontains=query) | Q(requests__icontains=query) | Q(cname__icontains=query)
                        | Q(finish__icontains=query) | Q(user_dept__icontains=query) | Q(
                            user_name__icontains=query)).order_by('-rg_date')
                else:
                    m_sheet = main_sheet.objects.all().order_by('-rg_date')

            # 한달이상 미처리건 조회
            over_date = main_sheet.objects.filter(rg_date__lte=date.today() - relativedelta(months=1)).exclude(
                finish='종료').order_by('-rg_date')

            # 페이징
            page = request.GET.get('page', '1')
            paginator = Paginator(m_sheet, 10)
            page_obj = paginator.get_page(page)
            print("insung GET main 페이징 끝")
            context = {'login_session': login_session, 'page_obj': page_obj, 'sort': sort, 'over_date': over_date,
                       'query': query, 'search_sort': search_sort, 'user_name': user_name, 'user_dept': user_dept,
                       'sdate': startdate, 'edate': enddate}

        else:
            sort = request.GET.get('sort', '')
            query = request.GET.get('q', '')
            if sort == 'rg_date':
                m_sheet = main_sheet.objects.filter(cname=login_session).order_by('-rg_date')
            elif sort == 'rp_date':
                m_sheet = main_sheet.objects.filter(cname=login_session).order_by('-rp_date')
            elif sort == 'main_title':
                m_sheet = main_sheet.objects.filter(cname=login_session).order_by('-main_title', '-rg_date')
            elif sort == 'cname':
                m_sheet = main_sheet.objects.filter(cname=login_session).order_by('cname', '-rg_date')
            elif sort == 'total_price':
                m_sheet = main_sheet.objects.filter(cname=login_session).order_by('total_price', '-rg_date')
            elif sort == 'finish':
                m_sheet = main_sheet.objects.filter(cname=login_session).order_by('finish', '-rg_date')
            elif sort == 'all':
                m_sheet = main_sheet.objects.filter(cname=login_session).order_by('-rg_date', 'finish')
            else:
                print("리스트 조회 겸 목록 조회")
                search_sort = request.GET.get('search_sort', '')
                startdate = request.GET.get('sdate', '')
                enddate = request.GET.get('edate', '')
                if search_sort == 'main_title':
                    m_sheet = main_sheet.objects.all().filter(main_title__icontains=query,
                                                              cname=login_session).order_by('-rg_date')
                elif search_sort == 'requests':
                    m_sheet = main_sheet.objects.all().filter(requests__icontains=query,
                                                              cname=login_session).order_by('-rg_date')
                elif search_sort == 'user_dept':
                    m_sheet = main_sheet.objects.all().filter(user_dept__icontains=query, cname=login_session).order_by(
                        '-rg_date')
                elif search_sort == 'user_name':
                    m_sheet = main_sheet.objects.all().filter(user_name__icontains=query, cname=login_session).order_by(
                        '-rg_date')
                elif search_sort == 'rg_date':
                    e_date = datetime.datetime.strptime(enddate, '%Y-%m-%d') + datetime.timedelta(hours=23, minutes=59,
                                                                                                  seconds=59)
                    m_sheet = main_sheet.objects.all().filter(rg_date__gte=startdate, rg_date__lte=e_date,
                                                              cname=login_session).order_by('-rg_date')
                elif search_sort == 'rp_date':
                    m_sheet = main_sheet.objects.all().filter(rp_date__range=[startdate, enddate],
                                                              cname=login_session).order_by('-rg_date')
                elif search_sort == 'end_date':
                    m_sheet = main_sheet.objects.all().filter(end_date__range=[startdate, enddate],
                                                              cname=login_session).order_by('-rg_date')
                elif search_sort == 'product_name':
                    s_sheet = sub_sheet.objects.all().filter(product_name__icontains=query,
                                                             cname=login_session).order_by('-rg_date')
                    v = []
                    for i in s_sheet:
                        sm = i.m_id_id
                        ms = main_sheet.objects.filter(id__iexact=sm)
                        v = v + list(ms)
                    # 중복제거 / 정렬 작업을 위한 조건문
                    m_sheet = reduce(lambda acc, cur: acc if cur in acc else acc + [cur], v, [])
                elif search_sort == 'all':
                    m_sheet = main_sheet.objects.filter(Q(requests__icontains=query) | Q(finish__icontains=query) |
                                                        Q(main_title__icontains=query) | Q(user_dept__icontains=query)
                                                        | Q(user_name__icontains=query), cname=login_session).order_by(
                        '-rg_date')
                else:
                    m_sheet = main_sheet.objects.filter(cname=login_session).order_by('-rg_date')

            page = request.GET.get('page', '1')
            paginator = Paginator(m_sheet, 10)
            page_obj = paginator.get_page(page)
            print("일반 GET main 페이징 끝")
            context = {'login_session': login_session, 'page_obj': page_obj, 'sort': sort, 'query': query,
                       'search_sort': search_sort, 'sdate': startdate, 'edate': enddate,
                       'user_name': user_name, 'user_dept': user_dept}

        print('메인 리스트 끝')
        return render(request, 'isscm/main_list.html', context)
    elif request.method == 'POST':
        print('post 확인 필요')
        return redirect('isscm/main_list')


# sub 입력
def sub_insert(request, pk):
    print('sub 입력 시작')
    login_session = request.session.get('login_session')
    user_dept = request.session.get('user_dept')
    user_name = request.session.get('user_name')
    detailView = get_object_or_404(main_sheet, id=pk)

    if request.method == 'GET':
        print('sub 겟 들어옴')
        context = {'login_session': login_session, 'user_dept': user_dept, 'user_name': user_name,
                   'detailView': detailView}
        print('겟 끝나 나감')
        return render(request, 'isscm/sub_insert.html', context)
    elif request.method == 'POST':
        print("sub post 입력 시작")
        s_sheet = sub_sheet()
        s_sheet.product_name = request.POST['product_name']
        s_sheet.quantity = request.POST['quantity'].replace(",", "")
        s_sheet.per_price = request.POST.get('per_price').replace(",", "")
        s_sheet.tax = request.POST.get('tax').replace(",", "")
        s_sheet.total_price = request.POST.get('total_price').replace(",", "")
        s_sheet.cname = request.POST['cname']
        s_sheet.m_id = main_sheet.objects.get(id=pk)
        s_sheet.m_title = request.POST['main_title']
        s_sheet.user_name = user_name

        s_sheet.save()

        sub_total = sub_sheet.objects.filter(m_id_id=pk).distinct().values(
            'm_id_id').aggregate(Sum('total_price'))
        detailView.total_price = sub_total['total_price__sum']
        detailView.save()

        print('sub post 입력 종료')
        # 중복 업로드 방지
        return HttpResponseRedirect(reverse('isscm:main_detail', args=[pk]))


def sub_modify(request, pk, mid):
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')
    user_dept = request.session.get('user_dept')
    detailView = get_object_or_404(main_sheet, id=mid)
    sub_detailView = get_object_or_404(sub_sheet, id=pk)
    if request.method == 'GET':
        print('get sub detail 뷰 시작')
        context = {'sub_detailView': sub_detailView, 'login_session': login_session, 'user_name': user_name,
                   'user_dept': user_dept}
        print("디테일 뷰 끝")
        return render(request, 'isscm/sub_modify.html', context)
    else:
        print("post sub detail 뷰 / 수정 시작")
        # 수정 내용 저장
        sub_detailView.product_name = request.POST['product_name']
        # 숫자 , 제거 후 db 저장
        sub_detailView.per_price = request.POST.get('per_price').replace(",", "")
        sub_detailView.quantity = request.POST.get('quantity').replace(",", "")
        sub_detailView.tax = request.POST.get('tax').replace(",", "")
        sub_detailView.total_price = request.POST.get('total_price').replace(",", "")
        sub_detailView.cname = request.POST['cname']
        print("수정 저장완료")
        sub_detailView.save()

        # distinct()로 중복 제거 후 total_price 값 합계 구함
        sub_total = sub_sheet.objects.filter(m_id_id=mid).distinct().values(
            'm_id_id').aggregate(Sum('total_price'))
        detailView.total_price = sub_total['total_price__sum']
        detailView.save()

        sub = sub_sheet.objects.all().filter(m_id_id=mid).order_by('rg_date')

        print("insung GET main 페이징 끝")
        context = {'sub_detailView': sub_detailView, 'login_session': login_session, 'user_name': user_name,
                   'user_dept': user_dept, 'sub': sub, 'detailView': detailView}
        return render(request, 'isscm/main_detail.html', context)


# sub 삭제
def sub_delete(request, pk, mid):
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')
    user_dept = request.session.get('user_dept')
    detailView = get_object_or_404(main_sheet, id=mid)
    sub_detailView = get_object_or_404(sub_sheet, id=pk)
    sub_detailView.delete()
    print('삭제완료')

    sub_total = sub_sheet.objects.filter(m_id_id=pk).distinct().values(
        'm_id_id').aggregate(Sum('total_price'))
    detailView.total_price = sub_total['total_price__sum']
    detailView.save()

    sub = sub_sheet.objects.all().filter(m_id_id=mid).order_by('-rg_date')
    context = {'login_session': login_session, 'user_name': user_name,
               'user_dept': user_dept, 'sub': sub, 'detailView': detailView}
    # 삭제 후 값 가지고 다시 돌아감
    return render(request, 'isscm/main_detail.html', context)


# main 엑셀 다운로드 openpyxl 사용
def main_excel_openpyxl(request):
    login_session = request.session.get('login_session')

    print("main 다운로드 시작")
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response["Content-Disposition"] = 'attachment; filename=' \
                                      + str(datetime.date.today()) + '_main.xlsx'

    # 엑셀 오픈
    wb = openpyxl.Workbook()
    # 엑셀 시트 활성화
    ws = wb.active
    ws.title = 'main'

    # 첫번째 열에 들어갈 컬럼명 설정
    columns = ['등록 일자', '마감 요청 일자', '종료일자', '견적명', '업체명', '요청 사항', '총 금액', '종료 여부', '담당 팀', '담당자']

    query = request.GET.get('q')
    print('que : ', query)
    search_sort = request.GET.get('search_sort', '')
    print('search : ', search_sort)
    if search_sort:
        startdate = request.GET.get('sdate', '')
        enddate = request.GET.get('edate', '')
        print('enddate : ', enddate)
        print('검색으로 다운로드')
        if login_session == 'insung':
            print('insung search 다운')
            if search_sort == 'main_title':
                rows = main_sheet.objects.all().filter(main_title__icontains=query)
            elif search_sort == 'requests':
                rows = main_sheet.objects.all().filter(requests__icontains=query)
            elif search_sort == 'cname':
                rows = main_sheet.objects.all().filter(cname__icontains=query)
            elif search_sort == 'finish':
                rows = main_sheet.objects.all().filter(finish__icontains=query)
            elif search_sort == 'user_dept':
                rows = main_sheet.objects.all().filter(user_dept__icontains=query)
            elif search_sort == 'user_name':
                rows = main_sheet.objects.all().filter(user_name__icontains=query)
            elif search_sort == 'product_name':
                s_sheet = sub_sheet.objects.all().filter(product_name__icontains=query).order_by(
                    '-rg_date')
                v = []
                for i in s_sheet:
                    sm = i.m_id_id
                    ms = main_sheet.objects.filter(id__exact=sm)
                    v = v + list(ms)
                rows = reduce(lambda acc, cur: acc if cur in acc else acc + [cur], v, [])
            elif search_sort == 'all':
                rows = main_sheet.objects.all().filter(
                    Q(requests__icontains=query) | Q(cname__icontains=query) | Q(finish__icontains=query) |
                    Q(main_title__icontains=query) | Q(user_dept__icontains=query) |
                    Q(user_name__icontains=query))
            elif search_sort == 'rg_date':
                e_date = datetime.datetime.strptime(enddate, '%Y-%m-%d') + datetime.timedelta(hours=23, minutes=59,
                                                                                              seconds=59)
                print('엑셀다운 등록일자', e_date)
                rows = main_sheet.objects.all().filter(rg_date__gte=startdate, rg_date__lte=e_date).order_by('-rg_date')
            elif search_sort == 'rp_date':
                rows = main_sheet.objects.all().filter(rp_date__range=[startdate, enddate]).order_by(
                    '-rg_date')
            elif search_sort == 'end_date':
                rows = main_sheet.objects.all().filter(end_date__range=[startdate, enddate]).order_by(
                    '-rg_date')
        else:
            print('일반 search 다운')
            if search_sort == 'main_title':
                rows = main_sheet.objects.all().filter(main_title__icontains=query,
                                                       cname=login_session)
            elif search_sort == 'requests':
                rows = main_sheet.objects.all().filter(requests__icontains=query,
                                                       cname=login_session)
            elif search_sort == 'finish':
                rows = main_sheet.objects.all().filter(finish__icontains=query, cname=login_session)
            elif search_sort == 'product_name':
                s_sheet = sub_sheet.objects.all().filter(product_name__icontains=query, cname=login_session).order_by(
                    '-rg_date')
                v = []
                for i in s_sheet:
                    sm = i.m_id_id
                    ms = main_sheet.objects.filter(id__exact=sm)
                    v = v + list(ms)
                rows = reduce(lambda acc, cur: acc if cur in acc else acc + [cur], v, [])
            elif search_sort == 'all':
                rows = main_sheet.objects.filter(Q(requests__icontains=query) | Q(finish__icontains=query) |
                                                 Q(main_title__icontains=query), cname=login_session)
            elif search_sort == 'rg_date':
                e_date = datetime.datetime.strptime(enddate, '%Y-%m-%d') + datetime.timedelta(hours=23, minutes=59,
                                                                                              seconds=59)
                print('엑셀다운 등록일자', e_date)
                rows = main_sheet.objects.all().filter(rg_date__gte=startdate, rg_date__lte=e_date,
                                                       cname=login_session).order_by('-rg_date')
            elif search_sort == 'rp_date':
                rows = main_sheet.objects.all().filter(rp_date__range=[startdate, enddate],
                                                       cname=login_session).order_by('-rg_date')
            elif search_sort == 'end_date':
                rows = main_sheet.objects.all().filter(end_date__range=[startdate, enddate],
                                                       cname=login_session).order_by('-rg_date')
    else:
        print('전체 다운로드')
        if login_session == 'insung':
            rows = main_sheet.objects.all()
        else:
            rows = main_sheet.objects.all().filter(cname=login_session)

    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title

    for mainrow in rows:
        row_num += 1

        row = [
            # 날짜 포멧 변경
            mainrow.rg_date.strftime('%Y-%m-%d'),
            mainrow.rp_date,
            mainrow.end_date,
            mainrow.main_title,
            mainrow.cname,
            mainrow.requests,
            mainrow.total_price,
            mainrow.finish,
            mainrow.user_dept,
            mainrow.user_name,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value

    wb.save(response)
    print("다운로드 끝")
    return response


# sub 엑셀 다운로드
def sub_excel(request):
    login_session = request.session.get('login_session')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response["Content-Disposition"] = 'attachment; filename=' \
                                      + str(datetime.date.today()) + '_sub.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'sub'

    columns = ['메인 ID', '메인 Title', '등록 일자', '업체명', '제품명', '수량', '개당 단가', '부가세', '총 금액']

    m_id = request.GET.get('id')
    print("이건 뭘까", login_session)
    rows = sub_sheet.objects.filter(m_id_id=m_id)

    # 첫번째 열: 설정한 컬럼명 순서대로 스타일 적용하여 생성
    print("다운 중간2")
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title
        print('여기까지')

    for subex in rows:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            subex.m_id,
            subex.m_title,
            subex.rg_date.strftime('%Y-%m-%d'),
            subex.cname,
            subex.product_name,
            subex.quantity,
            subex.per_price,
            subex.tax,
            subex.total_price
        ]
        print('여기까지2')

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            print(cell_value)
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            print('여기까지3')

    wb.save(response)
    print("다운로드 끝")
    return response


# 메인 파일 업로드
def main_uploadFile(request, pk):
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')

    if request.method == "POST":
        if request.FILES.get('uploadedFile') is not None:
            if get_object_or_404(main_sheet, id=pk):
                print("pk 왓나요", pk)
                # 템플릿에서 데이터 가져오기

                cname = login_session
                main_id = main_sheet.objects.get(id=pk)
                menu = request.POST["menu"]

                files = request.FILES.getlist('uploadedFile')

                for f in files:
                    fileTitle = f
                    uploadedFile = f
                    # DB에 저장
                    uploadfile = models.UploadFile(
                        cname=cname,
                        title=fileTitle,
                        uploadedFile=uploadedFile,
                        main_id=main_id,
                        menu=menu
                    )
                    uploadfile.save()
                return HttpResponseRedirect(reverse('isscm:main_uploadFile', args=[pk]))
    else:
        print("get 으로 왓나", pk)
        detailView = get_object_or_404(main_sheet, id=pk)
        uploadfile = UploadFile.objects.filter(main_id=pk)
        no = pk
        context = {'login_session': login_session, 'no': no, 'detailView': detailView, 'files': uploadfile,
                   'user_name': user_name}
        print("겟 다 나갓나")
        return render(request, "isscm/file_upload.html", context)

    uploadfile = UploadFile.objects.filter(main_id=pk)
    detailView = get_object_or_404(main_sheet, id=pk)

    return render(request, "isscm/file_upload.html", context={'user_name': user_name,
                                                              "files": uploadfile, "login_session": login_session,
                                                              'detailView': detailView})


# 파일 다운로드 (바로보기가 아닌 다운로드를 위해서 만듬)
def main_downloadfile(request, pk):
    upload_file = get_object_or_404(UploadFile, no=pk)
    file = upload_file.uploadedFile
    name = file.name
    response = HttpResponse(content_type=mimetypes.guess_type(name)[0] or 'application/octet-stream')
    response['Content-Disposition'] = f'attachment; filename={name}'

    # copyfileobj : 파일 복사하여 다운로드 진행
    shutil.copyfileobj(file, response)
    return response


# 견적 파일 삭제
def main_file_delete(request, pk):
    login_session = request.session.get('login_session')
    sheetfile = get_object_or_404(UploadFile, no=pk)
    page_no = sheetfile.main_id_id
    if sheetfile.cname == login_session or login_session == 'insung':
        sheetfile.delete()
        print('삭제완료')
        return redirect(f'/main_uploadFile/{page_no}')
    else:
        print("삭제 됨?")
        return redirect(f'/main_detail/{pk}')


# sub 보증기간 리스트
@login_required
def product_list(request):
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')
    user_dept = request.session.get('user_dept')
    print("제품관리 리스트 시작")
    global search_sort
    global startdate
    global enddate

    if request.method == 'GET':
        if login_session == 'insung':
            print('get insung 리스트 시작')
            sort = request.GET.get('sort', '')
            query = request.GET.get('q', '')
            if sort == 'rg_date':
                sub_list = sub_sheet.objects.all().order_by('-rg_date')
            elif sort == 'product_name':
                sub_list = sub_sheet.objects.all().order_by('-product_name', '-rg_date')
            elif sort == 'total_price':
                sub_list = sub_sheet.objects.all().order_by('-total_price', '-rg_date', )
            elif sort == 'cname':
                sub_list = sub_sheet.objects.all().order_by('-cname', '-rg_date')
            elif sort == 'm_id':
                sub_list = sub_sheet.objects.all().order_by('-m_id', '-rg_date')
            elif sort == 'm_title':
                sub_list = sub_sheet.objects.all().order_by('-m_title', '-rg_date')
            elif sort == 'user_name':
                sub_list = sub_sheet.objects.all().order_by('-user_name', '-rg_date')
            elif sort == 'all':
                sub_list = sub_sheet.objects.all().order_by('-rg_date')
            else:
                print("리스트 조회 겸 목록 조회")
                search_sort = request.GET.get('search_sort', '')
                startdate = request.GET.get('sdate', '')
                enddate = request.GET.get('edate', '')
                if search_sort == 'product_name':
                    sub_list = sub_sheet.objects.all().filter(product_name__icontains=query).order_by(
                        '-product_name',
                        '-rg_date')
                elif search_sort == 'cname':
                    sub_list = sub_sheet.objects.all().filter(cname__icontains=query).order_by('-cname',
                                                                                               '-rg_date')
                elif search_sort == 'm_title':
                    sub_list = sub_sheet.objects.all().filter(m_title__icontains=query).order_by('-m_title',
                                                                                                 '-rg_date')
                elif search_sort == 'user_name':
                    sub_list = sub_sheet.objects.all().filter(user_name__icontains=query).order_by('-user_name',
                                                                                                   '-rg_date')
                elif search_sort == 'serial':
                    sub_list = []
                    product_list = product_info.objects.filter(serial__icontains=query).values('s_id').distinct()
                    print('product_list : ', product_list)
                    for i in product_list:
                        sid = i['s_id']
                        v = sub_sheet.objects.filter(id__exact=sid).order_by('-rg_date')
                        sub_list = sub_list + list(v)
                        print('sub_list : ', sub_list)
                elif search_sort == 'all':
                    sub_list = sub_sheet.objects.all().filter(
                        Q(product_name__icontains=query) | Q(cname__icontains=query) |
                        Q(m_title__icontains=query) | Q(user_name__icontains=query)).order_by('-user_name',
                                                                                              '-rg_date')
                elif search_sort == 'rg_date':
                    e_date = datetime.datetime.strptime(enddate, '%Y-%m-%d') + datetime.timedelta(hours=23, minutes=59,
                                                                                                  seconds=59)
                    sub_list = sub_sheet.objects.all().filter(rg_date__gte=startdate, rg_date__lte=e_date).order_by(
                        '-rg_date')
                else:
                    print('그냥 조회')
                    sub_list = sub_sheet.objects.all().order_by('-rg_date')

            page = request.GET.get('page', '1')
            paginator = Paginator(sub_list, 10)
            page_obj = paginator.get_page(page)
            print("insung GET main 페이징 끝")

            context = {'login_session': login_session, 'user_dept': user_dept, 'user_name': user_name, 'sort': sort,
                       'query': query, 'search_sort': search_sort, 'page_obj': page_obj, 'sdate': startdate,
                       'edate': enddate}
        else:
            print('일반 리스트 시작')
            sort = request.GET.get('sort', '')
            query = request.GET.get('q', '')
            if sort == 'rg_date':
                sub_list = sub_sheet.objects.filter(cname=login_session).order_by('-rg_date')
            elif sort == 'product_name':
                sub_list = sub_sheet.objects.filter(cname=login_session).order_by('-product_name', '-rg_date')
            elif sort == 'total_price':
                sub_list = sub_sheet.objects.filter(cname=login_session).order_by('-total_price', '-rg_date', )
            elif sort == 'm_id':
                sub_list = sub_sheet.objects.filter(cname=login_session).order_by('-m_id', '-rg_date')
            elif sort == 'm_title':
                sub_list = sub_sheet.objects.filter(cname=login_session).order_by('-m_title', '-rg_date')
            elif sort == 'all':
                sub_list = sub_sheet.objects.filter(cname=login_session).order_by('-rg_date', 'id')
            else:
                print("리스트 조회 겸 목록 조회")
                search_sort = request.GET.get('search_sort', '')
                startdate = request.GET.get('sdate', '')
                enddate = request.GET.get('edate', '')
                if search_sort == 'product_name':
                    sub_list = sub_sheet.objects.all().filter(product_name__icontains=query,
                                                              cname=login_session).order_by('-product_name',
                                                                                            '-rg_date')
                elif search_sort == 'm_title':
                    sub_list = sub_sheet.objects.all().filter(m_title__icontains=query, cname=login_session).order_by(
                        '-m_title',
                        '-rg_date')
                elif search_sort == 'user_name':
                    sub_list = sub_sheet.objects.all().filter(user_name__icontains=query, cname=login_session).order_by(
                        '-user_name',
                        '-rg_date')
                elif search_sort == 'serial':
                    sub_list = []
                    product_list = product_info.objects.filter(serial__icontains=query, cname=login_session).values(
                        's_id').distinct()
                    print('product_list : ', product_list)
                    for i in product_list:
                        sid = i['s_id']
                        v = sub_sheet.objects.filter(id__exact=sid).order_by('-rg_date')
                        sub_list = sub_list + list(v)
                        print('sub_list : ', sub_list)
                        print(sub_list)
                elif search_sort == 'all':
                    sub_list = sub_sheet.objects.all().filter(
                        Q(product_name__icontains=query) | Q(m_title__icontains=query) |
                        Q(user_name__icontains=query) | Q(user_name__icontains=query), cname=login_session).order_by(
                        '-user_name',
                        '-rg_date')
                elif search_sort == 'rg_date':
                    e_date = datetime.datetime.strptime(enddate, '%Y-%m-%d') + datetime.timedelta(1)
                    sub_list = sub_sheet.objects.all().filter(rg_date__gte=startdate, rg_date__lte=e_date,
                                                              cname=login_session).order_by(
                        '-rg_date')
                else:
                    sub_list = sub_sheet.objects.filter(cname=login_session).order_by('-rg_date', 'id')
            # 페이징
            page = request.GET.get('page', '1')
            paginator = Paginator(sub_list, 10)
            page_obj = paginator.get_page(page)
            print("insung GET main 페이징 끝")

            context = {'page_obj': page_obj, 'login_session': login_session, 'user_name': user_name,
                       'user_dept': user_dept,
                       'search_sort': search_sort, 'query': query, 'sdate': startdate, 'edate': enddate}
        return render(request, 'isscm/product_list.html', context)
    elif request.method == 'POST':
        print('post 확인 필요')
        return redirect('isscm/product_list')


# 제품 상세 뷰 / 입력
def product_modify(request, pk):
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')
    user_dept = request.session.get('user_dept')

    sub_detailView = get_object_or_404(sub_sheet, id=pk)
    product_view = product_info.objects.filter(s_id_id=pk)
    sid_count = product_info.objects.filter(s_id_id=pk).count() + 1
    entercount = get_object_or_404(sub_sheet, id=pk)

    sort = request.GET.get('sort', '')
    query = request.GET.get('q', '')
    search_sort = request.GET.get('search_sort', '')
    if request.GET.get('sdate', '') is not None:
        startdate = request.GET.get('sdate', '')
    if request.GET.get('edate', '') is not None:
        enddate = request.GET.get('edate', '')
    page = request.GET.get('page', '')

    if request.method == 'GET':
        print('get sub detail 뷰 시작')

        context = {'sub_detailView': sub_detailView, 'product_view': product_view, 'login_session': login_session,
                   'user_name': user_name, 'user_dept': user_dept, 'sort': sort, 'query': query,
                   'search_sort': search_sort,
                   'sdate': startdate, 'edate': enddate, 'page': page}
        print("디테일 뷰 끝")
        return render(request, 'isscm/product_modify.html', context)
    else:
        print("post sub detail 뷰 / 수정 시작")
        # 수정 내용 저장
        product = product_info()
        product.product_name = request.POST['product_name']
        product.cname = request.POST['cname']
        product.serial = request.POST['serial']
        product.production_date = request.POST['production_date']
        product.release_date = request.POST['release_date']
        if request.POST['release_date'] != "":
            re_date = request.POST['release_date']
            # 출고일로부터 3년 뒤 계산하여 DB 저장
            warranty = datetime.datetime.strptime(re_date, '%Y-%m-%d') + relativedelta(years=3)
        product.warranty = warranty
        product.s_id_id = pk
        product.user_name = user_name

        try:
            ex_pm = Product_Management.objects.exclude(status='폐기')
            pm_modify = get_object_or_404(ex_pm, serial=request.POST['serial'])
            pm_modify.product_name = request.POST['product_name']
            pm_modify.current_location = request.POST['cname']
            pm_modify.status = "출고"
            pm_modify.serial = request.POST['serial']

            pm_modify.save()
            print("pm 까지 수정 저장완료")
        except:
            print("수정 저장완료")
            product.save()

        product.save()

        entercount.enter_quantity = sid_count
        entercount.save()

        product_view = product_info.objects.filter(s_id_id=pk)
        # context = {'sub_detailView': sub_detailView, 'product_view': product_view, 'login_session': login_session,
        #            'user_name': user_name, 'user_dept': user_dept}
        # return render(request, 'isscm/product_modify.html', context)
        # 새로고침 중복 입력 방지 HttpResponseRedirect  args 로 pk값 같이 넘김
        return HttpResponseRedirect(reverse('isscm:product_modify', args=[pk]))


def product_delete(request, pk, sid):
    product_view = get_object_or_404(product_info, id=pk)
    product_view.delete()

    sid_count = product_info.objects.filter(s_id_id=sid).count()
    print('sid : ', sid_count)
    entercount = get_object_or_404(sub_sheet, id=sid)
    entercount.enter_quantity = sid_count
    entercount.save()

    print('제품정보 삭제완료')
    return redirect(f'/product_modify/{sid}')


# sub list 엑셀 다운로드
def sub_list_excel_openpyxl(request):
    global rows
    login_session = request.session.get('login_session')

    print("sub list 다운로드 시작")
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response["Content-Disposition"] = 'attachment; filename=' \
                                      + str(datetime.date.today()) + '_sub_list.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'sub_list'

    # 첫번째 열에 들어갈 컬럼명 설정
    columns = ['등록 일자', '제품명', '수량', '출고 수량', '개당 단가', '부가세', '총 금액', '업체명', '메인번호', '메인 Title', '담당자']

    query = request.GET.get('q')
    print('que : ', query)
    search_sort = request.GET.get('search_sort', '')
    print('search : ', search_sort)
    if search_sort:
        print('검색으로 다운로드')
        startdate = request.GET.get('sdate', '')
        enddate = request.GET.get('edate', '')
        if login_session == 'insung':
            print('insung search 다운')
            if search_sort == 'product_name':
                rows = sub_sheet.objects.filter(product_name__icontains=query).order_by('-rg_date')
            elif search_sort == 'cname':
                rows = sub_sheet.objects.filter(cname__icontains=query).order_by('-rg_date')
            elif search_sort == 'serial':
                rows = list()
                product_list = product_info.objects.filter(serial__icontains=query).values('s_id').distinct()
                print('product_list : ', product_list)
                for i in product_list:
                    sid = i['s_id']
                    # 정확한 숫자로 필터 거르기위해 exact 사용
                    v = sub_sheet.objects.all().filter(id__exact=sid).order_by('-rg_date')
                    for w in v:
                        rows.append(w)
                print('rows : ', rows)
            elif search_sort == 'm_title':
                rows = sub_sheet.objects.filter(m_title__icontains=query).order_by('-rg_date')
            elif search_sort == 'user_name':
                rows = sub_sheet.objects.filter(user_name__icontains=query).order_by('-rg_date')
            elif search_sort == 'all':
                rows = sub_sheet.objects.all().filter(
                    Q(product_name__icontains=query) | Q(cname__icontains=query) | Q(serial__icontains=query) |
                    Q(m_title__icontains=query) | Q(user_name__icontains=query)).order_by('-rg_date')
            elif search_sort == 'rg_date':
                e_date = datetime.datetime.strptime(enddate, '%Y-%m-%d') + datetime.timedelta(hours=23, minutes=59,
                                                                                              seconds=59)

                rows = sub_sheet.objects.all().filter(rg_date__gte=startdate, rg_date__lte=e_date).order_by(
                    '-rg_date')
        else:
            print('일반 search 다운')
            if search_sort == 'product_name':
                rows = sub_sheet.objects.all().filter(product_name__icontains=query,
                                                      cname=login_session).order_by('-rg_date')
            elif search_sort == 'serial':
                rows = list()
                product_list = product_info.objects.filter(serial__icontains=query, cname=login_session).values(
                    's_id').distinct()
                print('product_list : ', product_list)
                for i in product_list:
                    sid = i['s_id']
                    # 정확한 숫자로 필터 거르기위해 exact 사용
                    v = sub_sheet.objects.all().filter(id__exact=sid).order_by('-rg_date')
                    for w in v:
                        rows.append(w)
            elif search_sort == 'm_title':
                rows = sub_sheet.objects.all().filter(m_title__icontains=query, cname=login_session).order_by(
                    '-rg_date')
            elif search_sort == 'user_name':
                rows = sub_sheet.objects.all().filter(user_name__icontains=query, cname=login_session).order_by(
                    '-rg_date')
            elif search_sort == 'all':
                rows = sub_sheet.objects.filter(Q(product_name__icontains=query) | Q(serial__icontains=query) |
                                                Q(m_title__icontains=query) | Q(user_name__icontains=query),
                                                cname=login_session).order_by('-rg_date')
            elif search_sort == 'rg_date':
                e_date = datetime.datetime.strptime(enddate, '%Y-%m-%d') + datetime.timedelta(hours=23, minutes=59,
                                                                                              seconds=59)

                rows = sub_sheet.objects.all().filter(rg_date__gte=startdate, rg_date__lte=e_date,
                                                      cname=login_session).order_by('-rg_date')
    else:
        print('전체 다운로드')
        if login_session == 'insung':
            rows = sub_sheet.objects.all().order_by('-rg_date')
        else:
            rows = sub_sheet.objects.all().filter(cname=login_session).order_by('-rg_date')

    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title

    for sublistrow in rows:
        row_num += 1
        row = [
            sublistrow.rg_date.strftime('%Y-%m-%d'),
            sublistrow.product_name,
            sublistrow.quantity,
            sublistrow.enter_quantity,
            sublistrow.per_price,
            sublistrow.tax,
            sublistrow.total_price,
            sublistrow.cname,
            sublistrow.m_id_id,
            sublistrow.m_title,
            sublistrow.user_name
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value

    wb.save(response)
    print("다운로드 끝")
    return response


# product info 엑셀 다운로드
def product_info_excel(request):
    login_session = request.session.get('login_session')

    print("main 다운로드 시작")
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response["Content-Disposition"] = 'attachment; filename=' \
                                      + str(datetime.date.today()) + '_product_info.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'product_info'

    columns = ['등록일자', '제품명', '업체명', '시리얼', '생산 일자', '출고 일자', '보증 만료일', 'SUB ID', '담당자']

    sub_id = request.GET.get('sid')
    if login_session == 'insung':
        print('인성 다운로드')
        rows = product_info.objects.filter(s_id_id=sub_id).order_by('-rg_date')
        print(rows.values())
    else:
        print('일반 다운로드')
        rows = product_info.objects.filter(s_id_id=sub_id, cname=login_session).order_by('-rg_date')
    print("다운 중간2")

    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title

    for plrow in rows:
        row_num += 1

        row = [
            plrow.rg_date.strftime('%Y-%m-%d'),
            plrow.product_name,
            plrow.cname,
            plrow.serial,
            plrow.production_date,
            plrow.release_date,
            plrow.warranty,
            plrow.s_id_id,
            plrow.user_name,
        ]
        print(row)

        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value

    wb.save(response)
    print("다운로드 끝")

    return response


# 제품명 DB
def product_db_insert(request):
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')
    user_dept = request.session.get('user_dept')

    if request.method == 'GET':

        context = {'login_session': login_session, 'user_name': user_name, 'user_dept': user_dept}
        return render(request, 'isscm/product_db_insert.html', context)
    else:
        product = ProductDb()
        product.center_code = request.POST['center_code']
        product.center = request.POST['center']
        product.warehouse_code = request.POST['warehouse_code']
        product.warehouse_name = request.POST['warehouse_name']
        product.product_code = request.POST['product_code']
        product.product_num = request.POST['product_num']
        product.scan_code = request.POST['scan_code']
        product.product_name = request.POST['product_name']
        product.account_code = request.POST['account_code']

        product.save()
        return HttpResponseRedirect(reverse('isscm:product_db_list'))


# 제품명 db 수정
def product_db_modify(request, pk):
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')
    user_dept = request.session.get('user_dept')
    # 해당 하는 row가 있다면 가져오고 없다면 404에러 발생
    detailview = get_object_or_404(ProductDb, no=pk)
    if request.method == 'GET':
        print('get 옴')
        context = {'detailview': detailview, 'login_session': login_session, 'user_name': user_name,
                   'user_dept': user_dept}
        return render(request, 'isscm/product_db_modify.html', context)
    else:
        print('포스트')
        detailview.center_code = request.POST['center_code']
        detailview.center = request.POST['center']
        detailview.warehouse_code = request.POST['warehouse_code']
        detailview.warehouse_name = request.POST['warehouse_name']
        detailview.product_code = request.POST['product_code']
        detailview.product_num = request.POST['product_num']
        detailview.scan_code = request.POST['scan_code']
        detailview.product_name = request.POST['product_name']
        detailview.account_code = request.POST['account_code']

        detailview.save()
        # 중복 db 저장 방지
        return HttpResponseRedirect(reverse('isscm:product_db_list'))


# 제품명 DB 리스트
@login_required
@login_ok
def product_db_list(request):
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')
    user_dept = request.session.get('user_dept')

    productlist = ProductDb.objects.all().order_by('no')

    context = {'productlist': productlist, 'login_session': login_session, 'user_name': user_name,
               'user_dept': user_dept}

    return render(request, 'isscm/product_db_list.html', context)


# 제품명 DB 삭제
def product_db_delete(request, pk):
    product_pick = get_object_or_404(ProductDb, no=pk)
    product_pick.delete()
    print('삭제 완료')

    return redirect('isscm:product_db_list')


# 공지사항 입력
@login_required
@login_ok
def notice_insert(request):
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')
    user_dept = request.session.get('user_dept')

    if request.method == 'GET':
        print('공지사항 입력 get')
        context = {'login_session': login_session, 'user_name': user_name, 'user_dept': user_dept}
        return render(request, 'isscm/notice_insert.html', context)
    else:
        noticeview = notice()
        noticeview.title = request.POST['title']
        noticeview.user_dept = request.POST['user_dept']
        noticeview.user_name = request.POST['user_name']
        noticeview.start_date = request.POST['start_date']
        noticeview.end_date = request.POST['end_date']
        noticeview.content = request.POST['content']

        noticeview.save()
        print('공지사항 입력 완료')
        return HttpResponseRedirect(reverse('isscm:index'))


# 공지사항 뷰
def notice_view(request, pk):
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')
    user_dept = request.session.get('user_dept')
    notice_view = get_object_or_404(notice, no=pk)

    if request.method == 'GET':
        print('공지사항 뷰')
        context = {'notice_view': notice_view, 'login_session': login_session, 'user_name': user_name,
                   'user_dept': user_dept}
        return render(request, 'isscm/notice_view.html', context)
    else:
        print('공지사항 수정')
        notice_view.title = request.POST['title']
        notice_view.user_dept = request.POST['user_dept']
        notice_view.user_name = request.POST['user_name']
        notice_view.start_date = request.POST['start_date']
        notice_view.end_date = request.POST['end_date']
        notice_view.content = request.POST['content']

        notice_view.save()
        print('공지사항 수정 완료')

        return HttpResponseRedirect(reverse('isscm:notice_view', args=[pk]))


# 공지사항 삭제
def notice_delete(request, pk):
    notice_pick = get_object_or_404(notice, no=pk)
    notice_pick.delete()
    print('삭제 완료')

    return redirect('isscm:index')


# 제품관리 DB 입력
@login_required
@login_ok
def pm_insert(request):
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')
    user_dept = request.session.get('user_dept')

    if request.method == 'GET':

        context = {'login_session': login_session, 'user_name': user_name, 'user_dept': user_dept}
        return render(request, 'isscm/pm_insert.html', context)
    else:
        pm = Product_Management()
        pm.product_name = request.POST['product_name']
        pm.serial = request.POST['serial']
        pm.current_location = request.POST['current_location']
        pm.status = request.POST['status']

        pm.save()
        return HttpResponseRedirect(reverse('isscm:pm_list'))


# 제품관리 DB 리스트
@login_required
@login_ok
def pm_list(request):
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')
    user_dept = request.session.get('user_dept')
    global search_sort
    global startdate
    global enddate

    if request.method == 'GET':
        if login_session == 'insung':
            print('get insung 리스트 시작')
            sort = request.GET.get('sort', '')
            query = request.GET.get('q', '')
            if sort == 'rg_date':
                pm_sheet = Product_Management.objects.all().order_by('-rg_date')
            elif sort == 'product_name':
                pm_sheet = Product_Management.objects.all().order_by('-product_name', '-rg_date')
            elif sort == 'serial':
                pm_sheet = Product_Management.objects.all().order_by('-serial', '-rg_date')
            elif sort == 'current_location':
                pm_sheet = Product_Management.objects.all().order_by('-current_location', '-rg_date', )
            elif sort == 'status':
                pm_sheet = Product_Management.objects.all().order_by('-status', '-rg_date')
            elif sort == 'all':
                pm_sheet = Product_Management.objects.all().order_by('-rg_date')
            else:
                print("리스트 조회 겸 목록 조회")
                search_sort = request.GET.get('search_sort', '')
                startdate = request.GET.get('sdate', '')
                enddate = request.GET.get('edate', '')
                if search_sort == 'product_name':
                    pm_sheet = Product_Management.objects.all().filter(product_name__icontains=query).order_by(
                        '-rg_date')
                elif search_sort == 'serial':
                    pm_sheet = Product_Management.objects.all().filter(serial__icontains=query).order_by('-rg_date')
                elif search_sort == 'current_location':
                    pm_sheet = Product_Management.objects.all().filter(current_location__icontains=query).order_by(
                        '-rg_date')
                elif search_sort == 'status':
                    pm_sheet = Product_Management.objects.all().filter(status__icontains=query).order_by('-rg_date')
                elif search_sort == 'rg_date':
                    e_date = datetime.datetime.strptime(enddate, '%Y-%m-%d') + datetime.timedelta(hours=23, minutes=59,
                                                                                                  seconds=59)
                    pm_sheet = Product_Management.objects.all().filter(rg_date__gte=startdate,
                                                                       rg_date__lte=e_date).order_by('-rg_date')
                elif search_sort == 'update_date':
                    e_date = datetime.datetime.strptime(enddate, '%Y-%m-%d') + datetime.timedelta(hours=23, minutes=59,
                                                                                                  seconds=59)
                    pm_sheet = Product_Management.objects.all().filter(update_date__gte=startdate,
                                                                       update_date__lte=e_date).order_by('-rg_date')
                elif search_sort == 'all':
                    pm_sheet = Product_Management.objects.filter(
                        Q(product_name__icontains=query) | Q(serial__icontains=query) | Q(
                            current_location__icontains=query)
                        | Q(status__icontains=query)).order_by('-rg_date')
                else:
                    pm_sheet = Product_Management.objects.all().order_by('-rg_date')

            # 페이징f
            page = request.GET.get('page', '1')
            paginator = Paginator(pm_sheet, 15)
            page_obj = paginator.get_page(page)
            print("insung GET main 페이징 끝")
            context = {'login_session': login_session, 'page_obj': page_obj, 'sort': sort,
                       'query': query, 'search_sort': search_sort, 'user_name': user_name, 'user_dept': user_dept,
                       'sdate': startdate, 'edate': enddate}

        else:
            sort = request.GET.get('sort', '')
            query = request.GET.get('q', '')
            if sort == 'rg_date':
                pm_sheet = Product_Management.objects.all().filter(cname=login_session).order_by('-rg_date')
            elif sort == 'product_name':
                pm_sheet = Product_Management.objects.all().filter(cname=login_session).order_by('-product_name',
                                                                                                 '-rg_date')
            elif sort == 'serial':
                pm_sheet = Product_Management.objects.all().filter(cname=login_session).order_by('-serial', '-rg_date')
            elif sort == 'current_location':
                pm_sheet = Product_Management.objects.all().filter(cname=login_session).order_by('-current_location',
                                                                                                 '-rg_date', )
            elif sort == 'status':
                pm_sheet = Product_Management.objects.all().filter(cname=login_session).order_by('-status', '-rg_date')
            elif sort == 'all':
                pm_sheet = Product_Management.objects.all().filter(cname=login_session).order_by('-rg_date')
            else:
                print("리스트 조회 겸 목록 조회")
                search_sort = request.GET.get('search_sort', '')
                startdate = request.GET.get('sdate', '')
                enddate = request.GET.get('edate', '')
                if search_sort == 'product_name':
                    pm_sheet = Product_Management.objects.all().filter(product_name__icontains=query,
                                                                       cname=login_session).order_by(
                        '-rg_date')
                elif search_sort == 'serial':
                    pm_sheet = Product_Management.objects.all().filter(serial__icontains=query,
                                                                       cname=login_session).order_by('-rg_date')
                elif search_sort == 'current_location':
                    pm_sheet = Product_Management.objects.all().filter(current_location__icontains=query,
                                                                       cname=login_session).order_by(
                        '-rg_date')
                elif search_sort == 'status':
                    pm_sheet = Product_Management.objects.all().filter(status__icontains=query,
                                                                       cname=login_session).order_by('-rg_date')
                elif search_sort == 'rg_date':
                    e_date = datetime.datetime.strptime(enddate, '%Y-%m-%d') + datetime.timedelta(hours=23, minutes=59,
                                                                                                  seconds=59)
                    pm_sheet = Product_Management.objects.all().filter(rg_date__gte=startdate,
                                                                       rg_date__lte=e_date,
                                                                       cname=login_session).order_by('-rg_date')
                elif search_sort == 'all':
                    pm_sheet = Product_Management.objects.filter(
                        Q(product_name__icontains=query) | Q(serial__icontains=query) | Q(
                            current_location__icontains=query)
                        | Q(status__icontains=query), cname=login_session).order_by('-rg_date')
                else:
                    pm_sheet = Product_Management.objects.all().filter(cname=login_session).order_by('-rg_date')

            page = request.GET.get('page', '1')
            paginator = Paginator(pm_sheet, 10)
            page_obj = paginator.get_page(page)
            print("일반 GET main 페이징 끝")
            context = {'login_session': login_session, 'page_obj': page_obj, 'sort': sort, 'query': query,
                       'search_sort': search_sort, 'sdate': startdate, 'edate': enddate,
                       'user_name': user_name, 'user_dept': user_dept}

        print('PM 리스트 끝')
        return render(request, 'isscm/pm_list.html', context)
    else:
        print("post로 리스트옴 확인요망")

        return redirect('isscm/pm_list')


def pm_modify(request, pk):
    login_session = request.session.get('login_session')
    user_name = request.session.get('user_name')
    user_dept = request.session.get('user_dept')
    detailView = get_object_or_404(Product_Management, no=pk)

    sort = request.GET.get('sort', '')
    query = request.GET.get('q', '')
    search_sort = request.GET.get('search_sort', '')
    if request.GET.get('sdate', '') is not None:
        startdate = request.GET.get('sdate', '')
    if request.GET.get('edate', '') is not None:
        enddate = request.GET.get('edate', '')
    page = request.GET.get('page', '')

    if request.method == 'GET':
        print('get sub detail 뷰 시작')
        context = {'detailView': detailView, 'login_session': login_session, 'user_name': user_name,
                   'user_dept': user_dept, 'sort': sort, 'query': query, 'search_sort': search_sort, 'sdate': startdate,
                   'edate': enddate, 'page': page}
        print("디테일 뷰 끝")
        return render(request, 'isscm/pm_modify.html', context)
    else:
        print("post sub detail 뷰 / 수정 시작")
        # 수정 내용 저장
        detailView.product_name = request.POST['product_name']
        detailView.serial = request.POST['serial']
        detailView.current_location = request.POST['current_location']
        detailView.status = request.POST['status']

        detailView.save()
        print("수정 저장 끝")
        return redirect('isscm:pm_list')


def pm_delete(request, pk):
    pm_view = get_object_or_404(Product_Management, no=pk)
    pm_view.delete()

    print('삭제완료')
    return redirect('isscm:pm_list')


# product_manage 엑셀 다운로드 openpyxl 사용
def pm_excel_openpyxl(request):
    login_session = request.session.get('login_session')

    print("pm_list 다운로드 시작")
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response["Content-Disposition"] = 'attachment; filename=' \
                                      + str(datetime.date.today()) + '_pm.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Product_manager'

    columns = ['등록 일자', '수정 일자', '제품명', '시리얼', '현재 위치', '상태']

    query = request.GET.get('q')
    print('que : ', query)
    search_sort = request.GET.get('search_sort', '')
    print('search : ', search_sort)
    if search_sort:
        print('검색으로 다운로드')
        startdate = request.GET.get('sdate', '')
        enddate = request.GET.get('edate', '')
        print('insung search 다운')
        if search_sort == 'product_name':
            rows = Product_Management.objects.filter(product_name__icontains=query).order_by('rg_date')
        elif search_sort == 'current_location':
            rows = Product_Management.objects.filter(current_location__icontains=query).order_by('rg_date')
        elif search_sort == 'serial':
            rows = Product_Management.objects.filter(serial__icontains=query).order_by('rg_date')
        elif search_sort == 'status':
            rows = Product_Management.objects.filter(status__icontains=query).order_by('rg_date')
        elif search_sort == 'rg_date':
            e_date = datetime.datetime.strptime(enddate, '%Y-%m-%d') + datetime.timedelta(hours=23, minutes=59,
                                                                                          seconds=59)
            rows = Product_Management.objects.filter(rg_date__gte=startdate, rg_date__lte=e_date).order_by(
                'rg_date')
        elif search_sort == 'update_date':
            d_date = datetime.datetime.strptime(enddate, '%Y-%m-%d') + datetime.timedelta(hours=23, minutes=59,
                                                                                          seconds=59)
            rows = Product_Management.objects.filter(rg_date__gte=startdate, rg_date__lte=d_date).order_by(
                'rg_date')
        elif search_sort == 'all':
            rows = Product_Management.objects.filter(
                Q(product_name__icontains=query) | Q(current_location__icontains=query) | Q(status__icontains=query)). \
                order_by('rg_date')
        else:
            rows = Product_Management.objects.all().order_by('rg_date')
    else:
        print('전체 다운로드')
        rows = Product_Management.objects.all().order_by('rg_date')

    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title

    for asrow in rows:
        row_num += 1

        row = [
            asrow.rg_date.strftime('%Y-%m-%d'),
            asrow.product_name,
            asrow.serial,
            asrow.current_location,
            asrow.status,
            asrow.update_date.strftime('%Y-%m-%d')
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value

    wb.save(response)
    print("다운로드 끝")
    return response


def pm_excel_upload(request):
    if request.method == "POST":
        print("엑셀 업로드 시작")
        try:
            files = request.FILES['uploadedFile']
            # data_only=True 수식이 아닌 값으로 읽어옴
            load_wb = openpyxl.load_workbook(files, data_only=True)
            # 시트 이름으로 불러오기
            # load_ws = load_wb['Sheet1']
            # 첫번째 시트 불러오기 - 해당 방법 채택
            load_ws = load_wb.worksheets[0]
            all_values = []
            for row in load_ws.rows:
                row_value = []
                for cell in row:
                    row_value.append(cell.value)
                all_values.append(row_value)

            for idx, val in enumerate(all_values):
                print('저장 진행중')
                # 첫번쩨(제목)줄 일때
                if idx == 0:
                    print('val : ', val[0])
                    print('val : ', val[1])
                    print('val : ', val[2])
                    # 엑셀 형식 체크 (첫번째의 제목 row) 제목 틀리면 안들어가고 빠져나옴
                    if val[0] != 'product_name' or val[1] != 'serial' or val[2] != 'current_location' or val[3] != 'status':
                        print("항목 틀림")
                        break
                else:
                    # 한줄 전체가 공란이면 안들어가고 브레이크로 빠져나옴
                    if val[0] == None and val[1] == None and val[2] == None and val[3] == None:
                        print("끝")
                        break
                    else:
                        print('idx : ', idx)
                        print('val0 : ', val[0])
                        print('val1 : ', val[1])
                        print('val2 : ', val[2])
                        print('val3 : ', val[3])
                        pm = Product_Management()
                        pm.product_name = val[0]
                        pm.serial = val[1]
                        pm.current_location = val[2]
                        pm.status = val[3]

                        pm.save()
                        print('저장 완료')
        except:
            return redirect('isscm:pm_list')

    return HttpResponseRedirect(reverse('isscm:pm_list'))
